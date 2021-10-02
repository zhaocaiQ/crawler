# 필요한 라이브러리를 로딩합니다.
from bs4 import BeautifulSoup
from selenium import webdriver
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
import requests
import urllib
import time
import os

print("="*80)
print(f'{"쿠팡 이미지크롤러 ver1.0": ^65}')
print("="*80)
# 검색어 입력받기
while True:
    search = input("검색어를 입력해주세요 > ")
    if search == "":
        print("검색할 검색어를 입력해주세요.")
    else:
        print("="*80)
        break

path = "C:\python_temp\chromedriver_win32\chromedriver.exe"
driver = webdriver.Chrome(path)
driver.get("https://www.coupang.com/np/search?component=&q="+search)
time.sleep(2)

# 시작시간
s_time = time.time()

html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')
# 이미지를 저장할 폴더를 생성합니다.
while True:
    f_dir = input("저장할 폴더를 지정하세요:(예: c:\\coupang_img\\) > ")
    if f_dir == "":
        print("저장할 폴더 경로를 입력해주세요.")
        continue
    elif f_dir[-1] != "\\":
        print("저장할 폴더 경로를 정확히 입력해주세요")
        continue
    else:
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        try:
            product = soup.find_all("li", "search-product")
        except AttributeError:
            print("정보가 없습니다.")
            print("="*80)
        else:
            nos = []
            titles = []
            prices = []
            stars = []
            reviews = []
            links = []
            no = 1
            for i in product:
                nos.append(no)
                title = i.find("div", "name").get_text()
                titles.append(title)
                price = i.find("strong", "price-value").get_text()
                prices.append(price)
                try:
                    star = i.find("span", "star").get_text()
                except:
                    star = '평점정보가 없습니다.'
                    stars.append(star)
                else:
                    stars.append(star)
                try:
                    review = i.find("span", "rating-total-count").get_text()
                except:
                    review = '리뷰가 없습니다.'
                    reviews.append(review)
                else:
                    reviews.append(review)
                link = i.find("a").attrs['href']
                links.append('https://www.coupang.com'+link)
                no += 1

            # 저장에 필요한 시간정보 추출하기
            now = time.localtime()
            f_name = "%04d-%02d-%02d" % (now.tm_year, now.tm_mon, now.tm_mday)
            # 생성한 폴더 안으로 이동
            f_result_dir = f_dir+f_name+'-'+search+'-'+"쿠팡"
            if os.path.isdir(f_result_dir):
                os.chdir(f_result_dir)
            else:
                os.makedirs(f_result_dir)
                os.chdir(f_result_dir)
            print("저장경로:", f_result_dir)
            print("="*80)

            # csv파일로 저장하기
            save_name = f_name+'-'+search+'-'+"쿠팡"+".xlsx"
            data = pd.DataFrame()
            data['순서'] = nos
            data['상품명'] = titles
            data['가격'] = prices
            data['평점'] = stars
            data['리뷰수'] = reviews
            data['링크'] = links
            if os.path.isfile(save_name):
                os.remove(save_name)
                data.to_excel(save_name, index=False)
            else:
                data.to_excel(save_name, index=False)

            # 이미지를 추출하여 저장합니다.
            count = 1
            img_srcs = []
            img_src = soup.find("ul", "search-product-list").find_all('img',
                                                                      attrs={"class": "search-product-wrap-img"})
            for img in img_src:
                img_src1 = img['src']
                if '.gif' in img_src1:
                    pass
                else:
                    img_srcs.append('https:'+img_src1)
                    count += 1
                try:
                    img_src2 = img['data-img-src']
                except:
                    pass
                else:
                    if '.gif' in img_src2:
                        pass
                    else:
                        img_srcs.append('https:'+img_src2)
                        count += 1
            file_no = 1
            for i in range(0, len(img_srcs)):
                img_name = f'{search}-{str(file_no)}.jpg'
                if os.path.isfile(img_name):
                    os.remove(img_name)
                    # img_src2[i]주소를 str(file_no)+'.jpg'라는 이름으로 변경
                    urllib.request.urlretrieve(img_srcs[i], img_name)
                else:
                    urllib.request.urlretrieve(img_srcs[i], img_name)
                time.sleep(0.5)
                print('%s번째 이미지 저장중입니다.=========' % file_no)
                # 엑셀파일 열기
                wb = openpyxl.load_workbook(save_name)
                ws = wb.active
                # 첫번째 시트 이름 변경하기
                names = []
                for name in wb.sheetnames:
                    names.append(name)
                for sheet in range(len(names)):
                    wb[names[sheet]].title = f_name+'-'+search+f'{sheet+1}'
                # G1에 '이미지' 텍스트삽입.굵게/테두리 생성
                ws['G1'] = '이미지'
                ws['G1'].alignment = Alignment(
                    horizontal='center', vertical='top')
                ws['G1'].font = Font(bold=True)
                ws['G1'].border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                # 엑셀파일에 이미지 삽입하기
                imgs = openpyxl.drawing.image.Image(
                    f'{search}-{str(file_no)}.jpg')
                # 셀크기 변경
                ws.column_dimensions['G'].width = 29
                ws.row_dimensions[i+2].height = 180
                # 이미지 추가하기
                imgs.anchor = 'G'+f'{i+2}'
                ws.add_image(imgs)
                # 엑셀파일 저장하기
                wb.save(save_name)
                file_no += 1
            # 작업시간 표시
            e_time = time.time()
            t_time = e_time-s_time
            print('='*80)
            print('출력에 걸린시간은 %s초 입니다' % round(t_time, 1))
            print('='*80)
            print(f"{save_name}으로 저장이 완료되었습니다.")
            print("쿠팡 이미지 크롤러를 종료합니다.")
            print("="*80)
        break
